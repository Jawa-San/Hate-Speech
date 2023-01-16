import openpyxl
from openpyxl import Workbook


wb = openpyxl.load_workbook('Dataset.xlsx')
ws = wb.create_sheet("Edit sheet")


with open('entries.txt') as f:
    lines = f.readlines()
# appending text and user input
x = 1
for i in range(1, len(lines), 2):
    ws.cell(x, 2, lines[i])
    print(lines[i])

    string = input("Speech type: ")
    ws.cell(x, 3, string)

    string = input("Is offensive: ")
    ws.cell(x, 4, string)

    string = input("Based on: ")
    ws.cell(x, 5, string)

    string = input("Target group: ")
    ws.cell(x, 6, string)

    string = input("Target sub group: ")
    ws.cell(x, 7, string)

    x += 1
# appending the urls
x = 1
for i in range(0, len(lines), 2):
    ws.cell(x, 1, lines[i])
    x += 1

wb.save('Dataset.xlsx')